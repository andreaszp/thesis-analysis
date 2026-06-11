"""
export.py
---------
Final step: write all results to a single formatted Excel file.

Color scheme (unified v2):
    Section titles:  dark navy (#1B2A4A) primary / dark grey (#2C3E50) secondary
    Column headers:  medium blue (#2980B9)
    Rows:            white / very light grey (#F2F3F4) alternating
    p-values:        grey (ns) / orange (p<.05) / green (p<.01) / dark green (p<.001)
    Mediation type:  dark green (full) / light green (partial)
    Section titles:  merged across full table width
"""

import logging
import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------
C_TITLE    = '1B2A4A'   # dark navy   — primary section titles
C_SUBTITLE = '2C3E50'   # dark grey   — secondary section titles
C_HEADER   = '2980B9'   # medium blue — column headers
C_ALT      = 'F2F3F4'   # light grey  — alternating rows
C_WHITE    = 'FFFFFF'
C_SIG_001  = '1E8449'   # dark green  — p<.001
C_SIG_01   = '27AE60'   # green       — p<.01
C_SIG_05   = 'F39C12'   # orange      — p<.05
C_NS       = 'E8E8E8'   # light grey  — ns
C_MED_FULL = '1E8449'   # dark green  — full mediation
C_MED_PART = 'A9DFBF'   # light green — partial mediation
C_MULTICOL = 'F39C12'   # orange      — multicollinearity warning
C_STEP     = 'D5D8DC'   # light grey  — step sub-headers

# Block colors for Sheet K (Cleaned Data)
BLOCK_COLORS = {
    'ai':          '1A5276',
    'eval':        '784212',
    'quality':     '1D6A39',
    'engage':      '4A235A',
    'demo':        '424949',
    'personality': '7B241C',
    'tone':        '1F618D',
    'default':     '1B4F72',
}

# Category colors for Sheet L (Variable Definitions)
CAT_COLORS = {
    'Experimental Design':       '1F618D',
    'Chatbot Evaluation':        '784212',
    'Perceived Manipulation':    '1A5276',
    'Competence to Judge':       '1A5276',
    'Moral Agency':              '1A5276',
    'Moral Patiency':            '1A5276',
    'Perceived Autonomy':        '1A5276',
    'Perceived Personality':     '7B241C',
    'Feedback Quality (GPT-4o)': '1D6A39',
    'Engagement':                '4A235A',
    'Demographics':              '424949',
}

# ---------------------------------------------------------------------------
# Labels
# ---------------------------------------------------------------------------
PREDICTOR_LABELS: dict = {
    'tone':             'Chatbot Tone (0=Formal / 1=Casual)',
    'PM_score':         'Perceived Manipulation (composite)',
    'PM1':              'PM — threat to freedom',
    'PM2':              'PM — decision override',
    'PM3':              'PM — manipulation attempt',
    'PM4':              'PM — pressure felt',
    'Comp_score':       'Competence (composite)',
    'Comp1':            'Competence — skills',
    'Comp2':            'Competence — morality',
    'Ind_score':        'Autonomy (composite)',
    'Ind1':             'Autonomy — AI plans & goals',
    'Ind2':             'Autonomy — AI self-control',
    'MA_score':         'Moral Agency (composite)',
    'MA1':              'Moral Agency — moral gravity AI→human',
    'MA2':              'Moral Agency — AI moral responsibility',
    'MP_score':         'Moral Patiency (composite)',
    'MP1':              'Moral Patiency — moral gravity human→AI',
    'MP2':              'Moral Patiency — AI right to consideration',
    'E1':               'Required effort',
    'E2':               'Engagement felt',
    'E3':               'Chatbot appreciation',
    'E4':               'Conversation utility',
    'E5':               'Reuse intention',
    'E6':               'Chatbot preference',
    'composite':        'Feedback composite score',
    'engagement_score': 'Engagement score (behavioural)',
    'emotions_mean':    'Emotional expression (mean)',
    'quantity_mean':    'Feedback quantity (mean)',
    'quality_mean':     'Feedback quality (mean)',
}

DV_LABELS: dict = {
    'composite':        'Feedback Composite Score',
    'engagement_score': 'Engagement Score (behavioural)',
    'emotions_mean':    'Emotional Expression',
    'E3':               'Chatbot Appreciation (E3)',
    'E4':               'Conversation Utility (E4)',
    'E5':               'Reuse Intention (E5)',
    'E6':               'Chatbot Preference (E6)',
}

IV_LABELS            = PREDICTOR_LABELS
DV_LABELS_REGRESSION = {
    'PM_score':  'Perceived Manipulation (composite)',
    'PM1':       'PM — threat to freedom',
    'PM2':       'PM — decision override',
    'PM3':       'PM — manipulation attempt',
    'PM4':       'PM — pressure felt',
    'Comp_score':'Competence (composite)',
    'Comp1':     'Competence — skills',
    'Comp2':     'Competence — morality',
    'Ind_score': 'Autonomy (composite)',
    'Ind1':      'Autonomy — AI plans & goals',
    'Ind2':      'Autonomy — AI self-control',
    'MA_score':  'Moral Agency (composite)',
    'MA1':       'Moral Agency — moral gravity AI→human',
    'MA2':       'Moral Agency — AI moral responsibility',
    'MP_score':  'Moral Patiency (composite)',
    'MP1':       'Moral Patiency — moral gravity human→AI',
    'MP2':       'Moral Patiency — AI right to consideration',
}

# ---------------------------------------------------------------------------
# Core style helpers
# ---------------------------------------------------------------------------
def _fill(hex_color: str) -> PatternFill:
    return PatternFill(
        start_color=hex_color, end_color=hex_color, fill_type='solid'
    )


def _write_value(cell, value) -> None:
    if isinstance(value, (np.integer,)):
        cell.value = int(value)
    elif isinstance(value, (np.floating,)):
        cell.value = float(value) if not np.isnan(value) else None
    elif isinstance(value, float) and np.isnan(value):
        cell.value = None
    elif isinstance(value, bool):
        cell.value = str(value)
    else:
        cell.value = value if value is not None else None


def _autofit(ws, min_w: int = 10, max_w: int = 38) -> None:
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, min(len(str(cell.value or '')), 50))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(
            max(max_len + 2, min_w), max_w
        )


def _write_section(ws, row: int, title: str,
                   n_cols: int, primary: bool = True) -> int:
    """
    Write a section title row merged across n_cols.
    Primary = dark navy. Secondary = dark grey.
    """
    color = C_TITLE if primary else C_SUBTITLE
    cell  = ws.cell(row=row, column=1)
    cell.value     = title
    cell.font      = Font(
        bold=True, size=11 if primary else 10, color='FFFFFF'
    )
    cell.fill      = _fill(color)
    cell.alignment = Alignment(vertical='center', wrap_text=False)
    ws.row_dimensions[row].height = 24 if primary else 20

    # Fill remaining cols with same color
    for col in range(2, n_cols + 1):
        ws.cell(row=row, column=col).fill = _fill(color)

    # Merge cells
    if n_cols > 1:
        try:
            ws.merge_cells(
                start_row=row, start_column=1,
                end_row=row,   end_column=n_cols
            )
        except Exception:
            pass

    # Thick bottom border
    ws.cell(row=row, column=1).border = Border(
        bottom=Side(style='medium', color='AAAAAA')
    )
    return row + 1


def _write_headers(ws, row: int, headers: list) -> int:
    """Write column header row in medium blue."""
    for col_idx, h in enumerate(headers, start=1):
        cell           = ws.cell(row=row, column=col_idx)
        cell.value     = h
        cell.fill      = _fill(C_HEADER)
        cell.font      = Font(bold=True, color='FFFFFF', size=9)
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )
        cell.border = Border(
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )
    ws.row_dimensions[row].height = 30
    return row + 1


def _write_df(ws, df: pd.DataFrame, start_row: int,
              color_p: bool = True,
              color_med: bool = False,
              n_cols_hint: int = None) -> int:
    """
    Write a DataFrame with alternating rows, p-value coloring,
    and optional mediation-type coloring.
    """
    if df is None or df.empty:
        cell       = ws.cell(row=start_row, column=1)
        cell.value = 'No results.'
        cell.font  = Font(italic=True, color='888888', size=9)
        return start_row + 2

    headers = list(df.columns)
    n_cols  = len(headers)

    start_row = _write_headers(ws, start_row, headers)

    p_names   = {'p', 'p-value', 'p_fdr', 'p_a', 'p_b', 'p_c',
                 'p_cprime', 'p_fchange', 'p (e3)', 'p (e4)', 'p (e5)',
                 'p (interaction)', 'pp2 p', 'p (f)'}
    sig_names = {'sig.', 'sig', 'sig. e3', 'sig. e4', 'sig. e5', 'pp2 sig.'}
    med_names = {'mediation_type', 'type'}

    for row_idx, (_, row) in enumerate(df.iterrows()):
        alt  = (row_idx % 2 == 0)
        base = _fill(C_ALT) if alt else _fill(C_WHITE)

        for col_idx, value in enumerate(row, start=1):
            cell           = ws.cell(row=start_row, column=col_idx)
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.fill      = base
            _write_value(cell, value)

            h = str(headers[col_idx - 1]).lower()

            # p-value coloring
            if color_p and h in p_names:
                try:
                    p = float(value)
                    if p < .001:
                        cell.fill = _fill(C_SIG_001)
                        cell.font = Font(size=9, bold=True, color='FFFFFF')
                    elif p < .01:
                        cell.fill = _fill(C_SIG_01)
                        cell.font = Font(size=9, bold=True, color='FFFFFF')
                    elif p < .05:
                        cell.fill = _fill(C_SIG_05)
                        cell.font = Font(size=9, bold=True)
                    else:
                        cell.fill = _fill(C_NS)
                except (ValueError, TypeError):
                    pass

            # Sig. coloring
            if h in sig_names:
                v = str(value)
                if '***' in v:
                    cell.fill = _fill(C_SIG_001)
                    cell.font = Font(size=9, bold=True, color='FFFFFF')
                elif '**' in v:
                    cell.fill = _fill(C_SIG_01)
                    cell.font = Font(size=9, bold=True, color='FFFFFF')
                elif '*' in v and 'ns' not in v:
                    cell.fill = _fill(C_SIG_05)
                    cell.font = Font(size=9, bold=True)
                elif v == 'ns':
                    cell.fill = _fill(C_NS)
                    cell.font = Font(size=9, color='666666')

            # Mediation type coloring
            if color_med and h in med_names:
                v = str(value)
                if 'Full' in v:
                    cell.fill = _fill(C_MED_FULL)
                    cell.font = Font(size=9, bold=True, color='FFFFFF')
                elif 'Partial' in v:
                    cell.fill = _fill(C_MED_PART)
                    cell.font = Font(size=9, bold=True, color='1A5276')

            # Multicollinearity warning
            if h == 'multicollinearity':
                if '⚠️' in str(value) or 'r > 0.80' in str(value):
                    cell.fill = _fill(C_MULTICOL)
                    cell.font = Font(size=9, bold=True)

        start_row += 1

    # Thick border after last data row
    for col in range(1, n_cols + 1):
        ws.cell(row=start_row - 1, column=col).border = Border(
            bottom=Side(style='medium', color='AAAAAA')
        )

    return start_row + 1


def _note(ws, row: int, text: str) -> int:
    """Write an italic note row."""
    cell           = ws.cell(row=row, column=1)
    cell.value     = text
    cell.font      = Font(italic=True, size=8.5, color='555555')
    cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 22
    return row + 1


# ---------------------------------------------------------------------------
# Sheet writers — all using unified style
# ---------------------------------------------------------------------------

def _write_sheet_toc(ws, df: pd.DataFrame) -> None:
    """Table of Contents."""
    ws.sheet_view.showGridLines = False
    if df is None or df.empty:
        return
    headers = list(df.columns)
    n_cols  = len(headers)

    _write_section(ws, 1, 'SoundFlow Thesis — Results Excel', n_cols)
    current_row = _write_headers(ws, 2, headers)

    for row_idx, (_, row) in enumerate(df.iterrows()):
        alt = (row_idx % 2 == 0)
        for col_idx, value in enumerate(row, start=1):
            cell       = ws.cell(row=current_row, column=col_idx)
            cell.fill  = _fill(C_ALT) if alt else _fill(C_WHITE)
            cell.font  = Font(size=10)
            cell.value = str(value) if value else ''
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        current_row += 1

    _autofit(ws, min_w=20, max_w=80)


def _write_sheet_l(ws, df: pd.DataFrame) -> None:
    """Variable Definitions — grouped by category."""
    categories = {
        'Experimental Design':       ['tone'],
        'Chatbot Evaluation':        ['E1','E2','E3','E4','E5','E6'],
        'Perceived Manipulation':    ['PM1','PM2','PM3','PM4','PM_score'],
        'Competence to Judge':       ['Comp1','Comp2','Comp_score'],
        'Moral Agency':              ['MA1','MA2','MA_score'],
        'Moral Patiency':            ['MP1','MP2','MP_score'],
        'Perceived Autonomy':        ['Ind1','Ind2','Ind_score'],
        'Perceived Personality':     ['PP1','PP2','PP3','PP4','PP5'],
        'Feedback Quality (GPT-4o)': [
            'quantity_run1/2/3','quantity_mean',
            'quality_run1/2/3','quality_mean',
            'emotions_run1/2/3','emotions_mean','composite',
        ],
        'Engagement': [
            'avg_words_per_turn','chat_duration_sec',
            'z_avg_words','z_duration','engagement_score',
            'conversation_completed','total_user_words','num_turns',
        ],
        'Demographics': ['age','gender','language'],
    }

    if df is None or df.empty:
        return

    headers  = list(df.columns)
    n_cols   = len(headers)
    curr_row = 1

    for cat_name, var_list in categories.items():
        color = CAT_COLORS.get(cat_name, C_TITLE)

        # Category title (merged, colored)
        cell       = ws.cell(row=curr_row, column=1)
        cell.value = cat_name
        cell.font  = Font(bold=True, size=10, color='FFFFFF')
        cell.fill  = _fill(color)
        cell.alignment = Alignment(vertical='center')
        ws.row_dimensions[curr_row].height = 20
        for col in range(2, n_cols + 1):
            ws.cell(row=curr_row, column=col).fill = _fill(color)
        try:
            ws.merge_cells(
                start_row=curr_row, start_column=1,
                end_row=curr_row,   end_column=n_cols
            )
        except Exception:
            pass
        curr_row += 1

        # Column headers
        curr_row = _write_headers(ws, curr_row, headers)

        # Data rows
        cat_df = df[df['Variable'].isin(var_list)].copy()
        for row_idx, (_, row) in enumerate(cat_df.iterrows()):
            alt = (row_idx % 2 == 0)
            for col_idx, value in enumerate(row, start=1):
                cell      = ws.cell(row=curr_row, column=col_idx)
                cell.fill = _fill(C_ALT) if alt else _fill(C_WHITE)
                cell.font = Font(size=9)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                _write_value(cell, value)
            curr_row += 1

        curr_row += 1  # blank row between categories

    _autofit(ws, min_w=15, max_w=55)


def _write_sheet_k(ws, df: pd.DataFrame) -> None:
    """Cleaned Data — color-coded headers by variable type."""
    def _col_color(col: str) -> str:
        if col.startswith(('PM', 'Comp', 'Ind', 'MA', 'MP')):
            return BLOCK_COLORS['ai']
        if col.startswith('PP'):
            return BLOCK_COLORS['personality']
        if col.startswith('E') and col[1:].isdigit():
            return BLOCK_COLORS['eval']
        if col in ['composite','quantity_mean','quality_mean','emotions_mean',
                   'quantity_run1','quantity_run2','quantity_run3',
                   'quality_run1','quality_run2','quality_run3',
                   'emotions_run1','emotions_run2','emotions_run3',
                   'quantity_sd','quality_sd','emotions_sd']:
            return BLOCK_COLORS['quality']
        if col in ['engagement_score','z_avg_words','z_duration',
                   'avg_words_per_turn','chat_duration_sec',
                   'num_turns','total_user_words','conversation_completed']:
            return BLOCK_COLORS['engage']
        if col in ['age','gender','language']:
            return BLOCK_COLORS['demo']
        if col in ['tone','tone_raw']:
            return BLOCK_COLORS['tone']
        return BLOCK_COLORS['default']

    if df is None or df.empty:
        ws.cell(row=1, column=1).value = 'No data.'
        return

    headers = list(df.columns)
    n_cols  = len(headers)

    for col_idx, h in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col_idx)
        cell.value     = h
        cell.fill      = _fill(_col_color(h))
        cell.font      = Font(bold=True, color='FFFFFF', size=9)
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )
    ws.row_dimensions[1].height = 35

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        alt = (row_idx % 2 == 0)
        for col_idx, value in enumerate(row, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx)
            cell.fill      = _fill(C_ALT) if alt else _fill(C_WHITE)
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            _write_value(cell, value)

    _autofit(ws, min_w=8, max_w=25)


def _write_sheet_a(ws, data: dict) -> None:
    """Correlations — Label X/Y, FDR recap + full."""
    recap = data.get('recap', pd.DataFrame())
    full  = data.get('full',  pd.DataFrame())

    def _clean(df):
        if df.empty or 'note' in df.columns:
            return df
        df = df.copy()
        return df.drop(
            columns=[c for c in ['Variable X','Variable Y'] if c in df.columns]
        )

    recap   = _clean(recap)
    full    = _clean(full)
    n_cols  = max(
        len(recap.columns) if not recap.empty else 1,
        len(full.columns)  if not full.empty  else 1,
    )
    row = 1
    row = _write_section(ws, row,
        'RECAP — Significant correlations only (FDR corrected)', n_cols)
    row = _write_df(ws, recap, row, color_p=True)
    row = _write_section(ws, row,
        'FULL RESULTS — All tested pairs', n_cols, primary=False)
    _write_df(ws, full, row, color_p=True)
    _autofit(ws)


def _write_sheet_b(ws, data: dict) -> None:
    """Effect of Tone — casual/formal, reordered columns."""

    def _reorder_cols(df):
        if df.empty or 'note' in df.columns:
            return df
        df = df.copy()
        # Rename friendly/professional → casual/formal everywhere
        rename_cols = {
            'M Friendly': 'M Casual', 'SD Friendly': 'SD Casual',
            'M Professional': 'M Formal', 'SD Professional': 'SD Formal',
            'N Friendly': 'N Casual', 'N Professional': 'N Formal',
            'Friendly': 'Casual', 'Professional': 'Formal',
        }
        df = df.rename(columns={
            c: rename_cols[c] for c in df.columns if c in rename_cols
        })
        # Replace values in all string columns
        df = df.replace({
            'Friendly': 'Casual', 'Professional': 'Formal',
            'friendly': 'casual', 'professional': 'formal',
            'FL_21': 'Casual (FL_21)', 'FL_22': 'Formal (FL_22)',
        })
        # Reorder columns: Variable | Label | p | d | Sig. | N Casual | M Casual | SD Casual | N Formal | M Formal | SD Formal | Δ | U/t
        preferred = [
            'Variable', 'Label',
            'p', "Cohen's d", 'Effect', 'Sig.',
            'N Casual', 'M Casual', 'SD Casual',
            'N Formal', 'M Formal', 'SD Formal',
            'Δ (C-F)', 'U / t',
        ]
        existing = [c for c in preferred if c in df.columns]
        remaining = [c for c in df.columns if c not in existing]
        return df[existing + remaining]

    blocks = [
        ('recap',       'RECAP — Significant results only',       True),
        ('personality', 'Perceived Personality',                   False),
        ('ai',          'AI Perceptions',                          False),
        ('eval',        'Chatbot Evaluation',                      False),
        ('quality',     'Quality & Engagement',                    False),
        ('h3',          'H3 — Paired t-test: Comp1 vs Comp2',     False),
    ]
    n_cols = max(
        len(v.columns) if isinstance(v, pd.DataFrame) and not v.empty else 1
        for v in data.values()
    ) if data else 12

    row = 1
    for key, title, primary in blocks:
        row = _write_section(ws, row, title, n_cols, primary=primary)
        df_block = _reorder_cols(data.get(key, pd.DataFrame()))
        row = _write_df(ws, df_block, row, color_p=True)
    _autofit(ws)


def _write_sheet_c(ws, data: dict) -> None:
    """AI Perception Regressions — no IV/DV label columns."""
    recap = data.get('recap', pd.DataFrame())
    full  = data.get('full',  pd.DataFrame())

    def _clean(df):
        if df.empty or 'note' in df.columns:
            return df
        df = df.copy()
        # Remove label columns
        drop = [c for c in ['IV Label', 'DV Label', 'Scale', 'IV_clean']
                if c in df.columns]
        return df.drop(columns=drop)

    recap  = _clean(recap)
    full   = _clean(full)
    n_cols = max(
        len(full.columns)  if not full.empty  else 1,
        len(recap.columns) if not recap.empty else 1,
    )
    row = 1
    row = _write_section(ws, row, 'RECAP — Significant results only', n_cols)
    row = _write_df(ws, recap, row, color_p=True)
    row = _write_section(ws, row,
        'FULL RESULTS — All a priori regressions', n_cols, primary=False)
    _write_df(ws, full, row, color_p=True)
    _autofit(ws)


def _write_hierarchical_sheet(ws, data: dict,
                               dvs: list, sheet_title: str) -> None:
    """Hierarchical regression with step sub-headers per DV."""
    full  = data.get('full',  pd.DataFrame())
    recap = data.get('recap', pd.DataFrame())

    def _clean(df):
        if df.empty or 'note' in df.columns:
            return df
        df = df.copy()
        # Remove Label column — keep Predictor
        return df.drop(
            columns=[c for c in ['Label', 'Block', 'Scale'] if c in df.columns]
        )

    full  = _clean(full)
    recap = _clean(recap)

    # Step labels — detect from Bloc column
    step_labels = {
        '1': 'Simple regression — Tone only',
        '2': 'Multiple regression — Tone + AI Perceptions',
        '3': 'Multiple regression — Tone + AI Perceptions + Engagement',
    }
    # For Sheet E the last step label differs
    if any(dv in dvs for dv in ['E3','E4','E5','E6']):
        step_labels['3'] = 'Multiple regression — Tone + AI Perceptions + Feedback Quality'

    n_cols = max(
        len(full.columns)  if not full.empty  else 1,
        len(recap.columns) if not recap.empty else 1,
    )

    row = 1
    row = _write_section(ws, row, 'RECAP — Significant results only', n_cols)
    row = _write_df(ws, recap, row, color_p=True)
    row = _write_section(ws, row, 'FULL RESULTS', n_cols)

    for dv in dvs:
        if full.empty or 'DV' not in full.columns:
            break
        df_dv = full[full['DV'] == dv].copy()
        if df_dv.empty:
            continue

        # DV title
        row = _write_section(
            ws, row, f'DV = {DV_LABELS.get(dv, dv)}',
            n_cols, primary=False
        )
        df_dv = df_dv.drop(columns=['DV'], errors='ignore')

        # Sub-headers per Bloc/step
        if 'Bloc' in df_dv.columns or 'Block' in df_dv.columns:
            bloc_col = 'Block' if 'Block' in df_dv.columns else 'Bloc'
            blocs = df_dv[bloc_col].unique()
            for bloc in sorted(blocs):
                df_bloc = df_dv[df_dv[bloc_col] == bloc].copy()
                df_bloc = df_bloc.drop(columns=[bloc_col], errors='ignore')
                step_label = step_labels.get(str(int(bloc)), f'Step {bloc}')

                # Step sub-header — light grey
                step_label = step_labels.get(str(bloc), f'Step {bloc}')
                for col in range(1, n_cols + 1):
                    cell      = ws.cell(row=row, column=col)
                    cell.fill = PatternFill(
                        start_color=C_STEP, end_color=C_STEP,
                        fill_type='solid'
                    )
                    cell.font      = Font(bold=True, size=9, color='2C3E50')
                    cell.alignment = Alignment(vertical='center')
                ws.cell(row=row, column=1).value = step_label
                try:
                    ws.merge_cells(
                        start_row=row, start_column=1,
                        end_row=row,   end_column=n_cols
                    )
                except Exception:
                    pass
                ws.row_dimensions[row].height = 18
                row += 1

                row = _write_df(ws, df_bloc, row, color_p=True)
        else:
            row = _write_df(ws, df_dv, row, color_p=True)

    _autofit(ws)

def _write_sheet_d(ws, data: dict) -> None:
    _write_hierarchical_sheet(
        ws, data,
        dvs=['composite', 'engagement_score', 'emotions_mean'],
        sheet_title='Predictors of Feedback Quality',
    )


def _write_sheet_e(ws, data: dict) -> None:
    _write_hierarchical_sheet(
        ws, data,
        dvs=['E3', 'E4', 'E5', 'E6'],
        sheet_title='Predictors of Chatbot Evaluation',
    )


def _write_sheet_f(ws, data: dict) -> None:
    """Mediation Analyses — reordered columns, casual/formal labels."""
    full  = data.get('full',  pd.DataFrame())
    recap = data.get('recap', pd.DataFrame())

    if full.empty or 'note' in full.columns:
        ws.cell(row=1, column=1).value = (
            'Run GPT scoring first to enable full mediation analyses.'
        )
        return

    def _reorder_med(df):
        if df.empty or 'note' in df.columns:
            return df
        df = df.copy()
        # Replace friendly/professional labels
        df = df.replace({
            'friendly': 'casual', 'professional': 'formal',
            'Friendly': 'Casual', 'Professional': 'Formal',
        })
        # Reorder: Series | IV | Mediator | DV | Mediation_type | Sig. | n | rest
        priority = ['Series', 'IV', 'Mediator', 'DV',
                    'Mediation_type', 'Sig.', 'n',
                    'a', 'p_a', 'b', 'p_b',
                    'c', 'p_c', 'c_prime', 'p_cprime',
                    'Indirect', 'CI_low', 'CI_high',
                    'Type', 'r_IV_M', 'r_IV_DV', 'r_M_DV',
                    'Multicollinearity']
        existing  = [c for c in priority if c in df.columns]
        remaining = [c for c in df.columns if c not in existing]
        return df[existing + remaining]

    full  = _reorder_med(full)
    recap = _reorder_med(recap)
    n_cols = len(full.columns)
    row    = 1

    row = _write_section(
        ws, row,
        'RECAP — Significant indirect effects (CI excludes 0)',
        n_cols
    )
    row = _write_df(ws, recap, row, color_p=False, color_med=True)

    for series_val, series_label in [
        ('1', 'SERIES 1 — Tone as IV'),
        ('2', 'SERIES 2 — AI Perceptions as IV'),
    ]:
        df_series = full[full['Series'] == series_val].copy()
        if df_series.empty:
            continue
        row = _write_section(ws, row, series_label, n_cols)

        for med in df_series['Mediator'].unique():
            df_med = df_series[df_series['Mediator'] == med].copy()
            if df_med.empty:
                continue
            med_label = PREDICTOR_LABELS.get(med, med)
            row = _write_section(
                ws, row, f'Mediator = {med_label}',
                n_cols, primary=False
            )
            row = _write_df(ws, df_med, row,
                            color_p=False, color_med=True)

    _autofit(ws)


def _write_sheet_g(ws, data) -> None:
    """GPT Scoring."""
    if isinstance(data, dict):
        scores  = data.get('scores',  pd.DataFrame())
        summary = data.get('summary', pd.DataFrame())
    else:
        scores  = data
        summary = pd.DataFrame()

    n_cols = max(
        len(scores.columns)  if not scores.empty  else 1,
        len(summary.columns) if not summary.empty else 1,
    )
    row = 1

    if not summary.empty:
        row = _write_section(ws, row, 'SUMMARY BY TONE CONDITION', n_cols)
        row = _write_df(ws, summary, row, color_p=False)

    row = _write_section(
        ws, row, 'FULL SCORES — All participants', n_cols, primary=False
    )
    _write_df(ws, scores, row, color_p=False)
    _autofit(ws)


def _write_sheet_h(ws, data: dict) -> None:
    """Word Frequencies."""
    if not data:
        ws.cell(row=1, column=1).value = 'No word frequency results.'
        return

    freq  = data.get('freq_table',  pd.DataFrame())
    tfidf = data.get('tfidf_table', pd.DataFrame())
    wcs   = data.get('wordcloud_paths', [])

    n_cols = max(
        len(freq.columns)  if not freq.empty  else 1,
        len(tfidf.columns) if not tfidf.empty else 1,
    )
    row = 1
    row = _write_section(
        ws, row, 'TABLE 1 — Word Frequencies (top 50 per condition)', n_cols
    )
    row = _write_df(ws, freq, row, color_p=False)
    row = _write_section(
        ws, row,
        'TABLE 2 — Delta TF-IDF '
        '(positive = more distinctive in Casual / '
        'negative = more distinctive in Formal)',
        n_cols
    )
    row = _write_df(ws, tfidf, row, color_p=False)

    if wcs:
        cell           = ws.cell(row=row, column=1)
        cell.value     = (
            'Wordcloud PNG files saved in outputs/wordclouds/: '
            + ', '.join(os.path.basename(p) for p in wcs)
        )
        cell.font      = Font(italic=True, size=9, color='555555')
        cell.alignment = Alignment(wrap_text=True)

    _autofit(ws)


def _write_sheet_i(ws, data: dict) -> None:
    """Demographics & Robustness."""
    n_cols   = 10
    sections = [
        ('DESCRIPTIVE STATISTICS',                   'descriptives',  True),
        ('RANDOMISATION CHECKS',                     'randomisation', False),
        ('EXCLUSION SUMMARY',                        'exclusions',    False),
        ('NORMALITY CHECKS (Shapiro-Wilk)',          'normality',     False),
        ('ANCOVA — Recap (significant only)',        'ancova_recap',  True),
        ('ANCOVA — Full results',                    'ancova',        False),
        ('INTERACTIONS — Recap (significant only)',  'inter_recap',   True),
        ('INTERACTIONS — Full results',              'interactions',  False),
    ]
    row = 1
    for title, key, primary in sections:
        row = _write_section(ws, row, title, n_cols, primary=primary)
        row = _write_df(ws, data.get(key, pd.DataFrame()), row, color_p=True)
    _autofit(ws)


def _write_sheet_supplementary(ws, supp: dict) -> None:
    """Supplementary Analyses sheet."""
    ws.sheet_view.showGridLines = False

    sections = [
        ('SECTION 1 — Supplementary Correlations',
         'correlations', True, True, False,
         '*** p<.001  ** p<.01  * p<.05  '
         'FDR correction not applied to these targeted pairs.'),
        ('SECTION 2 — Composite & Engagement by Tone × Language',
         'tone_language', True, False, False,
         'Tone × language interaction significant for engagement (F=4.20, p=.042) '
         'but not for composite quality.'),
        ('SECTION 3 — Emotional Expression by Tone Condition',
         'emotions_tone', True, False, False,
         'No significant difference by tone (t=1.530, p=.128).'),
        ('SECTION 4 — H3: Comp1 vs Comp2 by Tone',
         'h3_by_tone', True, False, False,
         'H3 effect is independent of tone condition (p=.353).'),
        ('SECTION 5 — Professional (PP2) vs Formal (PP5)',
         'pp2_vs_pp5', True, True, False,
         'Only PP2 (not PP5) is associated with positive evaluation and '
         'negatively with PM.'),
        ('SECTION 6 — Interaction Effects: PM × AI Perceptions → E3/E4/E5',
         'pm_interactions', True, True, False,
         'Significant interaction: PM × MP → E3 (β=.117, p=.047). '
         'All others ns.'),
        ('SECTION 7 — Simple Slopes: PM × MP → E3',
         'simple_slopes', False, False, False,
         'PM effect at M−1SD: β=−.525 | Mean: β=−.337 | M+1SD: β=−.149'),
        ('SECTION 8 — Moderation: Tone × PM → Outcomes',
         'tone_pm_moderation', True, True, False,
         'Marginal effect for reuse intention (p=.050). '
         'All other interactions ns.'),
        ('SECTION 9 — Mediation: Tone → PP1/PP3/PP4 → PM',
         'pp_pm_mediations', True, False, True,
         'All three personality mediations significant (CI excludes 0). '
         'Full mediation in each case.'),
        ('SECTION 10 — Serial Mediation: Tone → PM → Effort → Evaluation',
         'serial_mediations', True, False, True,
         'Not significant — PM effect is direct, not channelled via effort.'),
        ('SECTION 11 — E6 Regression with PP2',
         'e6_regression', True, True, False,
         'PP2: β=.240, p=.054 (quasi-significant). '
         'E6 remains largely unexplained.'),
        ('SECTION 11b — Model Comparison: E6 With vs Without PP2',
         'e6_model_comparison', False, False, False, None),
        ('SECTION 12 — Friendly+Professional vs Formal+Professional',
         'friendly_pro_comparison', True, True, False,
         'All differences non-significant (all p>.37). '
         'Consistent with correlational evidence but not confirmatory.'),
        ('SECTION PP-1 — Simple Regressions: Each PP Item → Composite',
         'pp1_simple_regressions', True, True, False,
         'OLS simple regression. Each PP item tested separately predicting composite quality.'),
        ('SECTION PP-2 — Multiple Regression: All PP Items → Composite',
         'pp2_multiple_regression', True, True, False,
         'OLS multiple regression. All PP items entered simultaneously predicting composite quality.'),
        ('SECTION PP-3 — Mediation: Tone → PP → Composite',
         'pp3_mediations', True, False, True,
         'Bootstrap 5,000 iterations. Tests whether personality perceptions mediate tone → composite.'),
        ('SECTION PP-4 — Serial Chain: Tone → PP → PM_score → Composite',
         'pp4_serial_composite', True, False, True,
         'Bootstrap 5,000 iterations. Full chain: tone → personality → manipulation → quality.'),
        ('SECTION PP-5 — Serial Chain: Tone → PP → PM_score → E3/E4/E5',
         'pp5_serial_eval', True, False, True,
         'Bootstrap 5,000 iterations. Full chain: tone → personality → manipulation → evaluation.'),
        ('SECTION PP-6 — Mediation: PP4 (Warm) → MP_score → E3',
         'pp6_pp4_mp_e3', True, False, True,
         'Bootstrap 5,000 iterations. Tests whether moral patiency mediates warmth → appreciation.'),
    ]

    # --- RECAP ---
    recap_rows = []
    for title, key, *_ in sections:
        if key not in supp:
            continue
        df_s = supp[key]
        if df_s.empty:
            continue
        sig_col = next(
            (c for c in df_s.columns if c.lower() in ('sig.', 'sig')), None
        )
        if sig_col:
            sig_df = df_s[
                df_s[sig_col].astype(str).str.contains(r'\*', na=False)
            ]
            for _, r in sig_df.iterrows():
                entry = {'Section': title.split(' — ')[-1]}
                for c in df_s.columns:
                    if c in ('Pair','Path','Interaction','DV','Predictor',
                             'β','β interaction','Indirect','CI_low',
                             'CI_high','r','p','Sig.','Type',
                             'Mediation_type','Interpretation'):
                        entry[c] = r.get(c, '')
                recap_rows.append(entry)

    n_cols_max = 12
    row        = 1

    row = _write_section(
        ws, row, 'RECAP — All Significant Supplementary Results',
        n_cols_max
    )
    if recap_rows:
        row = _write_df(
            ws, pd.DataFrame(recap_rows), row,
            color_p=True, color_med=True
        )
    else:
        ws.cell(row=row, column=1).value = 'No significant supplementary results.'
        row += 2

    row += 1

    # --- SECTIONS ---
    for title, key, primary, color_p, color_med, note in sections:
        n_cols = max(
            len(supp[key].columns)
            if key in supp and not supp[key].empty else 1,
            1
        )
        row = _write_section(
            ws, row, title, max(n_cols, n_cols_max), primary=primary
        )
        if note:
            row = _note(ws, row, note)
        if key in supp and not supp[key].empty:
            row = _write_df(
                ws, supp[key], row,
                color_p=color_p, color_med=color_med
            )
        else:
            ws.cell(row=row, column=1).value = 'No data.'
            row += 2
        row += 1

    _autofit(ws, min_w=10, max_w=40)


def _write_sheet_figures(ws, figure_paths: list) -> None:
    """Insert figure PNGs into the Figures sheet."""
    try:
        from openpyxl.drawing.image import Image as XLImage
    except ImportError:
        log.warning('openpyxl.drawing not available — skipping figure insertion.')
        return

    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2

    fig_labels = {
        'figure1':  'Figure 1 — Effect of Tone on Perceived Personality and PM',
        'figure2':  'Figure 2 — Full Path Diagram: Tone → Personality → PM → Outcomes',
        'figure3':  'Figure 3 — PM as Central Predictor of Chatbot Evaluation',
        'figure4':  'Figure 4 — Interaction: PM × Moral Patiency → Appreciation',
        'figure5':  'Figure 5 — Engagement Score and Quality by Tone × Language',
        'figure6':  'Figure 6 — Drivers of Reuse Intention',
        'figure7':  'Figure 7 — Comp1 (Skills) vs Comp2 (Morality) by Tone',
        'figure8':  'Figure 8 — Predictors of Feedback Quality (3-step regression)',
        'figure9':  'Figure 9 — Differential Predictors of E3/E4/E5',
        'figure10': 'Figure 10 — Professional (PP2) vs Formal (PP5): Correlations',
        'figure11': 'Figure 11 — Mediation: MA2 → MA1 → Appreciation',
        'figure12': 'Figure 12 — Competence-Autonomy Cluster and Moral Perceptions',
    }

    row = 1
    _write_section(ws, row, 'Thesis Figures — All 12 Figures', 15)
    row = 2

    for path in figure_paths:
        if not path or not os.path.exists(path):
            continue

        # Figure label from filename
        basename = os.path.basename(path).replace('.png', '')
        parts    = basename.split('_')
        fig_key  = parts[0] + parts[1] if len(parts) > 1 else basename
        label    = fig_labels.get(
            fig_key,
            basename.replace('_', ' ').title()
        )

        # Title row
        title_cell       = ws.cell(row=row, column=2)
        title_cell.value = label
        title_cell.font  = Font(bold=True, size=10, color='FFFFFF')
        title_cell.fill  = _fill(C_TITLE)
        title_cell.alignment = Alignment(vertical='center')
        ws.row_dimensions[row].height = 22
        try:
            ws.merge_cells(
                start_row=row, start_column=2,
                end_row=row,   end_column=15
            )
        except Exception:
            pass
        row += 1

        # Insert image
        try:
            img        = XLImage(path)
            img.width  = int(img.width  * 0.62)
            img.height = int(img.height * 0.62)
            ws.add_image(img, f'B{row}')
            n_rows = max(int(img.height / 14) + 2, 28)
            for r in range(row, row + n_rows):
                ws.row_dimensions[r].height = 14
            row += n_rows + 3
        except Exception as e:
            log.warning(f'Could not insert {path}: {e}')
            ws.cell(row=row, column=2).value = f'[Image not available: {path}]'
            row += 3


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------
def write_excel(
    results: dict,
    sheet_order: list,
    output_path: str,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    sheet_names = {
        'J': 'Table of Contents',
        'L': 'Variable Definitions',
        'K': 'Cleaned Data',
        'A': 'Correlations',
        'B': 'Effect of Tone',
        'C': 'AI Perception Regressions',
        'D': 'Predictors Feedback Quality',
        'E': 'Predictors Chatbot Evaluation',
        'F': 'Mediation Analyses',
        'G': 'GPT Scoring',
        'H': 'Word Frequencies',
        'I': 'Demographics Robustness',
    }

    # Main sheets
    for key in sheet_order:
        if key not in results:
            log.warning(f'Sheet {key}: no data — skipped.')
            continue

        name = sheet_names.get(key, key)
        ws   = wb.create_sheet(title=name)
        data = results[key]
        log.info(f'Writing: {name}')

        if key == 'J':
            _write_sheet_toc(ws, data)
        elif key == 'L':
            _write_sheet_l(ws, data)
        elif key == 'K':
            _write_sheet_k(ws, data)
        elif key == 'A':
            _write_sheet_a(ws, data if isinstance(data, dict) else {})
        elif key == 'B':
            _write_sheet_b(ws, data if isinstance(data, dict) else {})
        elif key == 'C':
            _write_sheet_c(ws, data if isinstance(data, dict) else {})
        elif key == 'D':
            _write_sheet_d(ws, data if isinstance(data, dict) else {})
        elif key == 'E':
            _write_sheet_e(ws, data if isinstance(data, dict) else {})
        elif key == 'F':
            _write_sheet_f(ws, data if isinstance(data, dict) else {})
        elif key == 'G':
            _write_sheet_g(ws, data)
        elif key == 'H':
            _write_sheet_h(ws, data if isinstance(data, dict) else {})
        elif key == 'I':
            _write_sheet_i(ws, data if isinstance(data, dict) else {})
        else:
            if isinstance(data, pd.DataFrame):
                _write_df(ws, data, start_row=1)
            else:
                ws.cell(row=1, column=1).value = f'No handler for {key}.'

        _autofit(ws)

    # Supplementary Analyses (after main sheets)
    if results.get('supp'):
        ws_supp = wb.create_sheet(title='Supplementary Analyses')
        _write_sheet_supplementary(ws_supp, results['supp'])
        log.info('Written: Supplementary Analyses')

    # Figures
    if results.get('figures'):
        ws_fig = wb.create_sheet(title='Figures')
        _write_sheet_figures(ws_fig, results['figures'])
        log.info('Written: Figures')

    try:
        wb.save(output_path)
        log.info(f'Saved: {output_path}')
    except Exception as e:
        log.error(f'Failed to save: {e}')
        raise
